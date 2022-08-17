import pandas as pd
import random
import torch
import time
import numpy as np
from gensim.models.word2vec import Word2Vec
from model import BatchProgramClassifier
from torch.autograd import Variable
from torch.utils.data import DataLoader
import os
import sys
from sklearn.metrics import roc_auc_score
from sklearn.metrics import accuracy_score
from sklearn.metrics import precision_score
from sklearn.metrics import f1_score
from sklearn.metrics import recall_score
import openpyxl

def get_batch(dataset, idx, bs):
    tmp = dataset.iloc[idx: idx+bs]
    data, labels = [], []
    for _, item in tmp.iterrows():
        data.append(item[1])
        # labels.append(item[2]-1)
        labels.append(item[2])
    return data, torch.LongTensor(labels)

def write_excel_xls_hypoth(path, sheet_name, value):
    if os.path.exists(path):
        data = openpyxl.load_workbook(path)
        table = data.create_sheet(sheet_name)
        table.title = sheet_name
        index = len(value)  # 获取需要写入数据的行数
        for i in range(0, index):
            for j in range(0, len(value[i])):
                table.cell(row=1 + i, column=j + 1, value=str(value[i][j]))  # 像表格中写入数据（对应的行和列）
        data.save(path)  # 保存工作簿
    else:
        index = len(value)
        workbook = openpyxl.Workbook()  # 新建工作簿（默认有一个sheet？）
        sheet = workbook.active  # 获得当前活跃的工作页，默认为第一个工作页
        sheet.title = sheet_name  # 给sheet页的title赋值
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))  # 行，列，值 这里是从1开始计数的
        workbook.save(path)  # 一定要保存
    print("xlsx格式表格写入数据成功！")

os.environ['CUDA_VISIBLE_DEVICES'] = '0'
res = [] 
# projects = ["hibernate", "lucene","ambari"]  # 1 13-15
# projects = ["lucene","argouml"]  # 2 16-30  还没完
# projects = ["ant", "argouml", "jackrabbit", "jenkins", "jmeter", "hibernate", "lucene","ambari"]
# projects = ["poi_3.1"]  
projects = ["hibernate"]  # 可新开个运行这些  "jmeter", "hibernate", "lucene" 
data = []
data_avg = []

# count = 10
count = 30
for project in projects:
    print('*********************', project, '*********************')
    data.append([project, '-', '-'])
    data_avg.append([project, '-', '-'])
    acc_sum = 0
    auc_sum = 0
    f1_sum = 0
    pre_sum = 0
    recall_sum = 0
    time_epoch = 0
    for h in range(30):
        
        print('*********************', '第',str(h), '轮', '*********************')
        root = 'data_cross/' + project +'/' +project+'_' + str(h) +'/'
        
        acc_cross_sum = 0
        auc_cross_sum = 0
        f1_cross_sum = 0
        pre_cross_sum = 0
        recall_cross_sum = 0
        # 10折交叉验证
        data_cross = []   # 保存交叉验证的中间结果
        for k in range(10):
            print('交叉验证----------', str(k))
            word2vec = Word2Vec.load(root+"embedding/node_w2v_128").wv
            embeddings = np.zeros((word2vec.wv.vectors.shape[0] + 1, word2vec.wv.vectors.shape[1]), dtype="float32")
            embeddings[:word2vec.wv.vectors.shape[0]] = word2vec.wv.vectors

            root_cross = root + str(k) + '/'

            train_data = pd.read_pickle(root_cross+'train/blocks.pkl')
            val_data = pd.read_pickle(root_cross + 'dev/blocks.pkl')
            test_data = pd.read_pickle(root_cross+'test/blocks.pkl')

            HIDDEN_DIM = 100
            # HIDDEN_DIM = 250  # hibernate
            # HIDDEN_DIM = 50   # ambari
            # HIDDEN_DIM = 200     # lucene  jackrabbit
            # HIDDEN_DIM = 240   # ant    1    10

            ENCODE_DIM = 128
            # LABELS = 104
            LABELS = 2
            # EPOCHS = 15
            EPOCHS = 10
            BATCH_SIZE = 50
            # BATCH_SIZE = 16
            USE_GPU = True
            # USE_GPU = False
            MAX_TOKENS = word2vec.wv.vectors.shape[0]
            EMBEDDING_DIM = word2vec.wv.vectors.shape[1]

            model = BatchProgramClassifier(EMBEDDING_DIM,HIDDEN_DIM,MAX_TOKENS+1,ENCODE_DIM,LABELS,BATCH_SIZE,
                                   USE_GPU, embeddings)
            if USE_GPU:
                model.cuda()

            parameters = model.parameters()
            optimizer = torch.optim.Adamax(parameters)
            # print(parameters)
            # print(list(parameters))
            loss_function = torch.nn.CrossEntropyLoss()

            train_loss_ = []
            val_loss_ = []
            train_acc_ = []
            val_acc_ = []
            best_acc = 0.0
            best_loss =1.0 
            print('Start training...')
            # training procedure
            best_model = model
            for epoch in range(EPOCHS):
                start_time = time.time()
                
                total_acc = 0.0
                total_loss = 0.0
                total = 0.0
                i = 0
                while i < len(train_data):
                    batch = get_batch(train_data, i, BATCH_SIZE)
                    i += BATCH_SIZE
                    train_inputs, train_labels = batch
                # print("train_inputs = ", train_inputs)
                # print("train_labels = ", train_labels)
                    if USE_GPU:
                        train_inputs, train_labels = train_inputs, train_labels.cuda()

                    model.zero_grad()
                    model.batch_size = len(train_labels)
                    model.hidden = model.init_hidden()
                    output = model(train_inputs)
                    # print("output = ", output)
                    loss = loss_function(output, Variable(train_labels))
                    loss.backward()
                    optimizer.step()

                    # calc training acc
                    _, predicted = torch.max(output.data, 1)
                    total_acc += (predicted == train_labels).sum()
                    total += len(train_labels)
                    total_loss += loss.item()*len(train_inputs)

                train_loss_.append(total_loss / total)
                train_acc_.append(total_acc.item() / total)
                # validation epoch
                total_acc = 0.0
                total_loss = 0.0
                total = 0.0
                i = 0
                while i < len(val_data):
                    batch = get_batch(val_data, i, BATCH_SIZE)
                    i += BATCH_SIZE
                    val_inputs, val_labels = batch
                    if USE_GPU:
                        val_inputs, val_labels = val_inputs, val_labels.cuda()

                    model.batch_size = len(val_labels)
                    model.hidden = model.init_hidden()
                    output = model(val_inputs)

                    loss = loss_function(output, Variable(val_labels))

                # calc valing acc
                    _, predicted = torch.max(output.data, 1)
                    total_acc += (predicted == val_labels).sum()
                    total += len(val_labels)
                    total_loss += loss.item()*len(val_inputs)
                val_loss_.append(total_loss / total)
                val_acc_.append(total_acc.item() / total)
                end_time = time.time()
                if total_acc/total >= best_acc:
                    best_model = model
                # if val_loss_[epoch]<best_loss:
                #     best_loss = val_loss_[epoch]
                #     best_model = model
                print('[Epoch: %3d/%3d] Training Loss: %.4f, Validation Loss: %.4f,'
                ' Training Acc: %.3f, Validation Acc: %.3f, Time Cost: %.3f s'
                % (epoch + 1, EPOCHS, train_loss_[epoch], val_loss_[epoch],
                    train_acc_[epoch], val_acc_[epoch], end_time - start_time))
                time_epoch = end_time - start_time

            i = 0
            model = best_model
            y_preds = []
            y_tests = []
        
            print(project,'--------------------')
            while i < len(test_data):
                batch = get_batch(test_data, i, BATCH_SIZE)
                i += BATCH_SIZE
                test_inputs, test_labels = batch
                if USE_GPU:
                    test_inputs, test_labels = test_inputs, test_labels.cuda()
                y_tests.append(test_labels)
                model.batch_size = len(test_labels)
                model.hidden = model.init_hidden()
                output = model(test_inputs)

                loss = loss_function(output, Variable(test_labels))

                _, predicted = torch.max(output.data, 1)
                y_preds.append(predicted)

                total_acc += (predicted == test_labels).sum()
                total += len(test_labels)
                total_loss += loss.item() * len(test_inputs)
        
            print("Testing results(Acc):", total_acc.item() / total)
            # print(y_preds)

            y_pred = []
            for y in y_preds:
                y_pred = y_pred + y.cpu().detach().numpy().tolist()
            y_pred = np.array(y_pred)

            y_test = []
            for y in y_tests:
                y_test = y_test + y.cpu().detach().numpy().tolist()
            y_test = np.array(y_test)
    
            acc = accuracy_score(y_test, y_pred)
            auc = roc_auc_score(y_test, y_pred)
            f1 = f1_score(y_test, y_pred)
            pre = precision_score(y_test, y_pred)
            recall = recall_score(y_test, y_pred)
            data_cross.append([k, acc, auc, f1, pre, recall])

            acc_cross_sum += acc
            auc_cross_sum += auc
            f1_cross_sum += f1
            pre_cross_sum += pre
            recall_cross_sum += recall

        # 每轮10折交叉验证结果
        acc_cross_avg = acc_cross_sum / 10
        auc_cross_avg = auc_cross_sum / 10
        f1_cross_avg = f1_cross_sum / 10
        pre_cross_avg = pre_cross_sum / 10
        recall_cross_avg = recall_cross_sum / 10

        # 保存交叉验证的中间结果和平均值
        data_cross.append(['avg', acc_cross_avg, auc_cross_avg, f1_cross_avg, pre_cross_avg, recall_cross_avg])
        dirs = 'result_cross-hibernate/'+project+'/'
        if not os.path.exists(dirs):
            os.makedirs(dirs)
        # path = dirs+project[:len(project)-1]+'_'+str(h)+'_cross.xls'
        path = dirs+project+'_'+str(h)+'_cross.xlsx'
        write_excel_xls_hypoth(path, 'cross-'+str(h), data_cross)
 

        data.append([acc_cross_avg, auc_cross_avg ,f1_cross_avg ,pre_cross_avg, recall_cross_avg])
        acc_sum += acc_cross_avg
        auc_sum += auc_cross_avg
        f1_sum += f1_cross_avg
        pre_sum += pre_cross_avg
        recall_sum += recall_cross_avg

        # dirs = 'result_cross/'+project
        # if not os.path.exists(dirs):
        #     os.makedirs(dirs)
        # path = dirs+project[:len(project)-1]+'_'+str(h)+'.xls'
        # write_excel_xls(path,project[:len(project)-1]+'_'+str(h) , data)
    
    dirs = 'result_cross-hibernate/'+project +'/'
    if not os.path.exists(dirs):
        os.makedirs(dirs)
    # path = dirs + project[:len(project)-1]+'.xls'
    path = dirs+project+'_'+str(h)+'.xlsx'
    write_excel_xls_hypoth(path, project+'_'+str(h), data)
    data = []
    data_avg.append([acc_sum/count, auc_sum/count, f1_sum/count, pre_sum/count, recall_sum/count, time_epoch])


    path = dirs + project+'_average'+'.xlsx'
    write_excel_xls_hypoth(path, 'average', data_avg)
    data_avg = []
    
    


