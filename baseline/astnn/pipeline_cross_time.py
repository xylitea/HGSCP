import pandas as pd
import os
import sys
import copy
import random
sys.setrecursionlimit(10000)
import time
import openpyxl

def write_excel_xls_hypoth(path, sheet_name, value):
    if os.path.exists(path):
        data = openpyxl.load_workbook(path)
        table = data.create_sheet(sheet_name)
        table.title = sheet_name
        index = len(value)  # 获取需要写入数据的行数
        for i in range(0, index):
            for j in range(0, len(value[i])):
                table.cell(row=1 + i, column=j + 1, value=str(value[i][j]))  # 像表格中写入数据（对应的行和列）
        data.save(path)  # 保存工作簿
    else:
        index = len(value)
        workbook = openpyxl.Workbook()  # 新建工作簿（默认有一个sheet？）
        sheet = workbook.active  # 获得当前活跃的工作页，默认为第一个工作页
        sheet.title = sheet_name  # 给sheet页的title赋值
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))  # 行，列，值 这里是从1开始计数的
        workbook.save(path)  # 一定要保存
    print("xlsx格式表格写入数据成功！")

data_time = []

'''
先对tree划分为语句tree，再进行交叉验证数据划分
'''

class Pipeline:
    def __init__(self,  ratio, root):
        self.ratio = ratio
        self.root = root
        self.sources = None
        self.train_file_path = None
        self.dev_file_path = None
        self.test_file_path = None
        self.size = None

    # parse source code
    def parse_source(self, output_file, option):
        path = self.root+output_file
        if os.path.exists(path) and option == 'existing':
            source = pd.read_pickle(path)
        else:
            # from pycparser import c_parser
            # parser = c_parser.CParser()
            import javalang
            def parse_program(func):
                tree = javalang.parse.parse(func)
                return tree

            source = pd.read_pickle(self.root+'programs.pkl')
            source.columns = ['id', 'code', 'label']
            source['code'] = source['code'].apply(parse_program)

            source.to_pickle(path)
        self.sources = source
        return source

    # construct dictionary and train word embedding
    def dictionary_and_embedding(self, input_file, size):
        self.size = size
        # if not input_file:
        #     input_file = self.train_file_path
        # trees = pd.read_pickle(input_file)
        trees = copy.deepcopy(self.sources)
        if not os.path.exists(self.root+'/embedding'):
            os.mkdir(self.root+'/embedding')
        from prepare_data import get_sequences

        def trans_to_sequences(ast):
            sequence = []
            get_sequences(ast, sequence)
            return sequence
        corpus = trees['code'].apply(trans_to_sequences)
        str_corpus = [' '.join(c) for c in corpus]
        trees['code'] = pd.Series(str_corpus)
        trees.to_csv(self.root+'/programs_ns.tsv')

        from gensim.models.word2vec import Word2Vec
        w2v = Word2Vec(corpus, size=size, workers=16, sg=1, min_count=3)
        w2v.save(self.root+'/embedding/node_w2v_' + str(size))

    # generate block sequences with index representations
    def generate_block_seqs(self):
        from prepare_data import get_blocks as func
        from gensim.models.word2vec import Word2Vec

        word2vec = Word2Vec.load(self.root+'/embedding/node_w2v_' + str(self.size)).wv
        vocab = word2vec.vocab
        max_token = word2vec.syn0.shape[0]

        def tree_to_index(node):
            token = node.token
            result = [vocab[token].index if token in vocab else max_token]
            children = node.children
            for child in children:
                result.append(tree_to_index(child))
            return result

        def trans2seq(r):
            blocks = []
            func(r, blocks)
            tree = []
            for b in blocks:
                btree = tree_to_index(b)
                tree.append(btree)
            return tree
        trees = self.sources
        trees['code'] = trees['code'].apply(trans2seq)
        trees.to_pickle(self.root+'/blocks.pkl')

    # split data for training, developing and testing
    def split_data(self):

        def check_or_create(path):
            if not os.path.exists(path):
                os.mkdir(path)

        def save_flod_data(i, part, data):  # i: 第几轮， part：训练、测试、验证 ， data：数据
            data_folder = self.root + '/' + str(i)
            check_or_create(data_folder)
            test_folder = data_folder +'/'+part
            check_or_create(test_folder)
            test_path = test_folder + '/blocks.pkl'
            data.to_pickle(test_path)

        data = pd.read_pickle(self.root+'/blocks.pkl')
        data_num = len(data)
        k = 10   # 10折交叉验证
        ratio = data_num // k
        data = data.sample(frac=1, random_state=random.randint(1,1000))

        for i in range(k):
            if i== k-1:
                test = data.iloc[ratio*i : ratio*(i+1)]
                dev = data.iloc[:ratio]
                train = data.iloc[ratio:ratio*i]
            else:
                test = data.iloc[ratio*i : ratio*(i+1)]
                dev = data.iloc[ratio*(i+1):ratio*(i+2)]
                train = data.iloc[:ratio*i].append(data.iloc[ratio*(i+2):])
             
            save_flod_data(i, 'test', test)
            save_flod_data(i, 'dev', dev)
            save_flod_data(i, 'train', train)
            
    # run for processing data to train
    def run(self):
        print('parse source code...')
        process_start = time.time()
        self.parse_source(output_file='ast.pkl',option='existing')
        process_end = time.time()
        print('train word embedding...')
        feature_start = time.time()
        self.dictionary_and_embedding(None,128)
        print('generate block sequences...')
        self.generate_block_seqs()
        feature_end = time.time()

        data_time.append([project, process_end-process_start, feature_end-feature_start])
        time_path = 'result_cross2/feature_time/'+project+'.xlsx'
        write_excel_xls_hypoth(time_path, 'Sheet1', data_time)

        print('end')


# projects = ['ant', 'argouml', 'hibernate', 'jenkins', 'jmeter', 'lucene']
# projects = ["ambari"]
projects = ['ambari', 'ant', 'argouml', 'hibernate', 'poi_3.1', 'jenkins', 'jmeter', 'lucene']

for project in projects:
    for i in range(1):
        ppl = Pipeline('7:1:2', 'data_cross_time/' + project + '/' + project + '_'+str(i)+ '/')
        ppl.run()


# ant
# ppl = Pipeline('3:1:1', 'data_cross/ant/ant_2/')
# ppl.run()


